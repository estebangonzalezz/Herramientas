#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Unificador de hojas Excel
· Tipos de planilla: 'lp' (por defecto) y 'simbiu'
· Para 'simbiu' se exige XML y se agrega la columna 'texto'
Autor: Esteban · Jun-2025 · MIT
"""

from __future__ import annotations

import io
import json
import os
import re
import unicodedata
import xml.etree.ElementTree as ET
from datetime import timedelta, datetime
from pathlib import Path
from typing import Dict, List, Tuple
from uuid import uuid4
from zipfile import BadZipFile

import openpyxl
import pandas as pd
from flask import (
    Flask,
    flash,
    redirect,
    render_template_string,
    request,
    send_file,
    abort,
    session,
    url_for,
)
from flask_session import Session
from functools import wraps

# ────────────────────── Configuración Flask ──────────────────────────
app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET", "cambia_esta_clave_supersecreta")

app.config["SESSION_TYPE"] = "filesystem"
app.config["SESSION_FILE_DIR"] = Path("flask_session").resolve()
app.config["SESSION_PERMANENT"] = True
Session(app)

app.permanent_session_lifetime = timedelta(hours=3)

MAPPINGS_DIR = Path("mappings")
MAPPINGS_DIR.mkdir(exist_ok=True)

PROCESSED_DIR = Path("processed")
PROCESSED_DIR.mkdir(exist_ok=True, parents=True)
HISTORY_FILE = PROCESSED_DIR / "history.json"

# ───────────────────────── Utilidades base ───────────────────────────
def canonicalize(text: str) -> str:
    """Minúsculas sin tildes ni símbolos; separa con '_'."""
    text = text.strip().lower()
    text = "".join(
        c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c)
    )
    text = re.sub(r"[^a-z0-9]", "_", text)
    return re.sub(r"__+", "_", text).strip("_")


def slugify(name: str) -> str:
    return canonicalize(name) or "default"


def workbook_columns(wb: openpyxl.Workbook) -> Dict[str, List[str]]:
    """Devuelve los encabezados de cada hoja."""
    return {
        ws.title: [
            c.value or f"col_{i+1}" for i, c in enumerate(next(ws.iter_rows(max_row=1)))
        ]
        for ws in wb.worksheets
    }


def list_mappings() -> List[str]:
    """Lista los tipos de planilla disponibles."""
    defaults = {"lp", "simbiu"}
    found = {p.stem for p in MAPPINGS_DIR.glob("*.json")}
    return sorted(defaults | found)


def save_mapping(name: str, mapping: Dict[str, str]) -> None:
    (MAPPINGS_DIR / f"{slugify(name)}.json").write_text(
        json.dumps(mapping, ensure_ascii=False, indent=2)
    )


def load_mapping(name: str) -> Dict[str, str]:
    p = MAPPINGS_DIR / f"{slugify(name)}.json"
    return json.loads(p.read_text()) if p.exists() else {}


def delete_mapping(name: str) -> None:
    """Elimina el archivo JSON correspondiente al mapeo."""
    p = MAPPINGS_DIR / f"{slugify(name)}.json"
    if p.exists():
        p.unlink()


def clone_mapping(src: str, dest: str) -> None:
    """Copia un mapeo existente con otro nombre."""
    src_p = MAPPINGS_DIR / f"{slugify(src)}.json"
    dest_p = MAPPINGS_DIR / f"{slugify(dest)}.json"
    if not src_p.exists():
        raise FileNotFoundError(src)
    if dest_p.exists():
        raise FileExistsError(dest)
    dest_p.write_bytes(src_p.read_bytes())


def load_history() -> List[Dict[str, str]]:
    if HISTORY_FILE.exists():
        try:
            return json.loads(HISTORY_FILE.read_text())
        except json.JSONDecodeError:
            return []
    return []


def add_history(mapping: str, filename: str) -> None:
    records = load_history()
    records.append({
        "mapping": mapping,
        "fecha": datetime.now().isoformat(timespec="seconds"),
        "archivo": filename,
    })
    HISTORY_FILE.write_text(json.dumps(records, ensure_ascii=False, indent=2))

# ───────────────────────── Texto (flujo simbiu) ──────────────────────
CODE_RE = re.compile(r"/(?:index/1|VerNoticia)/(\d+)")


def build_text_df(xml_bytes: bytes) -> pd.DataFrame:
    """Extrae código y texto completo del XML."""
    root = ET.parse(io.BytesIO(xml_bytes)).getroot()
    rows = []
    for n in root.findall(".//noticia"):
        url = n.findtext("Url_Noticia", "")
        m = CODE_RE.search(url)
        if m:
            rows.append(
                {
                    "codigo": m.group(1),
                    "texto": (n.findtext("FullText", "") or "").strip(),
                }
            )
    return pd.DataFrame(rows).drop_duplicates("codigo")


def add_text_column(df: pd.DataFrame, xml_bytes: bytes) -> pd.DataFrame:
    """Une el texto del XML a la planilla (impresos / digitales)."""
    if not xml_bytes:
        return df

    # localizar la columna con el enlace (case-insensitive)
    col_url = next(
        (c for c in df.columns if canonicalize(c) == "url_noticia"), None
    )
    if col_url is None:
        return df

    text_df = build_text_df(xml_bytes)
    df["codigo_link"] = df[col_url].astype(str).str.extract(CODE_RE, expand=False)
    df = df.merge(text_df, left_on="codigo_link", right_on="codigo", how="left")
    df.drop(columns=["codigo_link", "codigo"], inplace=True)
    return df

# ───────────────────────── Unificación de hojas ──────────────────────
def unify_workbook(xlsx_bytes: bytes, mapping: Dict[str, str]) -> pd.DataFrame:
    """
    Une todas las hojas del Excel normalizando encabezados.
    Devuelve DataFrame manteniendo los hipervínculos activos.
    """
    wb_in = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes), read_only=True, data_only=False
    )
    all_rows: List[List] = []
    headers_final: List[str] = []

    # ---- determinar encabezados finales --------------------------------
    for ws in wb_in.worksheets:
        for orig in next(ws.iter_rows(max_row=1, values_only=True)):
            orig = orig or ""
            dest = mapping.get(orig, mapping.get(canonicalize(orig), canonicalize(orig)))
            if dest not in headers_final:
                headers_final.append(dest)

    # ---- recorrer filas y re-mapear ------------------------------------
    for ws in wb_in.worksheets:
        header_src = [c or "" for c in next(ws.iter_rows(max_row=1, values_only=True))]
        idx2dest = {
            idx: mapping.get(orig, mapping.get(canonicalize(orig), canonicalize(orig)))
            for idx, orig in enumerate(header_src)
        }
        for row in ws.iter_rows(min_row=2, values_only=False):
            new_row = [""] * len(headers_final)
            for idx, cell in enumerate(row):
                dest = idx2dest.get(idx)
                if not dest:
                    continue
                if cell.hyperlink:
                    value = f'=HYPERLINK("{cell.hyperlink.target}", "{cell.value}")'
                elif isinstance(cell.value, str) and re.match(r"https?://", cell.value):
                    value = f'=HYPERLINK("{cell.value}", "{cell.value}")'
                else:
                    value = cell.value
                new_row[headers_final.index(dest)] = value
            all_rows.append(new_row)

    df = pd.DataFrame(all_rows, columns=headers_final)
    return df


def unify_files(files: List[Tuple[str, bytes]], mapping: Dict[str, str]) -> pd.DataFrame:
    """Une varias planillas y agrega la columna ``archivo_origen``."""
    merged = []
    for name, data in files:
        df = unify_workbook(data, mapping)
        df["archivo_origen"] = Path(name).stem
        merged.append(df)
    return pd.concat(merged, ignore_index=True) if merged else pd.DataFrame()

# ────────────────────────────── Rutas ────────────────────────────────

def login_required(view):
    """Redirige a /login si el usuario no está autenticado."""
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        user = request.form.get("username")
        pwd = request.form.get("password")
        if user == "AnalisisLP" and pwd == "AnalisisLP2025":
            session.permanent = True
            session["user"] = "AnalisisLP"
            return redirect(url_for("home"))
        error = "Credenciales incorrectas"
    return render_template_string(
        """
        <!doctype html>
        <title>Iniciar sesión</title>
        <link rel=stylesheet href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css">
        <div class="container py-5" style="max-width:400px">
          <h1 class="mb-4">Iniciar sesión</h1>
          {% if error %}<div class="alert alert-danger">{{ error }}</div>{% endif %}
          <form method="POST">
            <div class="mb-3">
              <label class="form-label">Usuario</label>
              <input class="form-control" name="username" required>
            </div>
            <div class="mb-3">
              <label class="form-label">Contraseña</label>
              <input class="form-control" type="password" name="password" required>
            </div>
            <button class="btn btn-primary">Entrar</button>
          </form>
        </div>
        """,
        error=error,
    )


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))
@app.route("/", methods=["GET", "POST"])
@login_required
def home():
    """
    Menú inicial: seleccionar tipo de planilla, subir Excel (.xlsx)
    y, si corresponde, XML.
    """
    mappings = list_mappings()

    if request.method == "POST":
        mapping_choice = slugify(request.form.get("mapping_choice", "lp"))
        files_xlsx = request.files.getlist("file_xlsx")
        f_xml = request.files.get("file_xml")

        # Validaciones
        if not files_xlsx or any(not f.filename.lower().endswith(".xlsx") for f in files_xlsx):
            flash("Sube un archivo Excel (.xlsx) válido.", "danger")
            return redirect(url_for("home"))

        if mapping_choice == "simbiu":
            if not (f_xml and f_xml.filename.lower().endswith(".xml")):
                flash("Para planillas Simbiu debes subir también el XML.", "danger")
                return redirect(url_for("home"))
            session["file_xml"] = f_xml.read()

        # Guardar en sesión
        session["files_xlsx"] = [(f.filename, f.read()) for f in files_xlsx]
        session["file_xlsx_name"] = Path(files_xlsx[0].filename).stem
        session["mapping_name"] = mapping_choice

        wb = openpyxl.load_workbook(io.BytesIO(session["files_xlsx"][0][1]), read_only=True)
        session["cols_per_sheet"] = workbook_columns(wb)

        return redirect(url_for("mapping"))

    selected = slugify(request.args.get("mapping", "lp"))
    return render_template_string(TPL_HOME, mappings=mappings, selected=selected)


@app.route("/mapping", methods=["GET", "POST"])
@login_required
def mapping():
    cols_per_sheet = session.get("cols_per_sheet")
    mapping_name = session.get("mapping_name", "lp")

    if not cols_per_sheet:
        return redirect(url_for("home"))

    if request.method == "POST":
        action = request.form.get("action", "unify")

        # Equivalencias capturadas (ignoramos el botón)
        mapping = {
            k: v.strip()
            for k, v in request.form.items()
            if k != "action" and v.strip()
        }
        save_mapping(mapping_name, mapping)

        if action == "save":
            flash("Configuración guardada", "success")
            return redirect(url_for("mapping"))

        try:
            # Unificación
            files_xlsx = session.get("files_xlsx", [])
            if len(files_xlsx) > 1:
                merged = unify_files(files_xlsx, mapping)
            else:
                merged = (
                    unify_workbook(files_xlsx[0][1], mapping)
                    if files_xlsx
                    else pd.DataFrame()
                )

            # Añadir texto si es Simbiu
            if mapping_name == "simbiu":
                merged = add_text_column(merged, session.get("file_xml", b""))

            out = io.BytesIO()
            merged.to_excel(out, index=False)
            out.seek(0)

            filename = f"{uuid4().hex}.xlsx"
            (PROCESSED_DIR / filename).write_bytes(out.getvalue())
            add_history(mapping_name, filename)

            out.seek(0)

            base = session.get("file_xlsx_name", slugify(mapping_name))
            session.clear()  # limpiamos todo
            return send_file(
                out,
                as_attachment=True,
                download_name=f"Unificado_{base}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except (BadZipFile, ValueError, Exception):
            flash("Ocurrió un error al procesar los archivos", "danger")
            app.logger.exception("Error al unificar archivos")
            return redirect(url_for("mapping"))

    return render_template_string(
        TPL_MAPPING,
        mapping_name=mapping_name,
        cols_per_sheet=cols_per_sheet,
        existing_mapping=load_mapping(mapping_name),
        canonicalize=canonicalize,
    )


@app.route("/mappings", methods=["GET", "POST"])
@login_required
def mappings_admin():
    """Administración básica de los tipos de planilla."""
    mappings = list_mappings()

    if request.method == "POST":
        action = request.form.get("action")

        if action == "create":
            new_name = request.form.get("new_name", "").strip()
            slug = slugify(new_name)
            if not slug:
                flash("Debes indicar un nombre", "danger")
            elif slug in [slugify(m) for m in mappings]:
                flash("Ese tipo ya existe", "danger")
            else:
                save_mapping(slug, {})
                flash("Tipo creado", "success")
                return redirect(url_for("home", mapping=slug))

        elif action == "clone":
            src = slugify(request.form.get("clone_src", ""))
            dest = request.form.get("clone_dest", "").strip()
            slug = slugify(dest)
            if not slug:
                flash("Debes indicar el nombre destino", "danger")
            else:
                try:
                    clone_mapping(src, slug)
                    flash("Tipo clonado", "success")
                    return redirect(url_for("home", mapping=slug))
                except FileExistsError:
                    flash("Ya existe un tipo con ese nombre", "danger")
                except FileNotFoundError:
                    flash("Tipo origen no encontrado", "danger")

        elif action == "delete":
            name = request.form.get("delete_name", "").strip()
            if name in {"lp", "simbiu"}:
                flash("No se puede borrar ese tipo", "danger")
            else:
                delete_mapping(name)
                flash("Tipo eliminado", "success")

        return redirect(url_for("mappings_admin"))

    return render_template_string(TPL_MAPPINGS, mappings=mappings)


@app.route("/processed/<path:filename>")
@login_required
def download_processed(filename: str):
    p = PROCESSED_DIR / filename
    if not p.exists():
        return abort(404)
    return send_file(p, as_attachment=True, download_name=filename)


@app.route("/historial")
@login_required
def historial():
    records = load_history()[::-1]
    return render_template_string(TPL_HISTORIAL, records=records)

# ────────────────────────── Plantillas HTML ──────────────────────────
TPL_HOME = """
<!doctype html>
<title>Unificar Excel · Menú</title>
<link rel="stylesheet"
 href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css">

<div class="container py-5">
  <h1 class="mb-4">Unificador de Hojas Excel</h1>
  <p><a href="{{ url_for('mappings_admin') }}">Administrar tipos</a></p>
  <p><a href="{{ url_for('historial') }}">Historial de planillas</a></p>
  {% if session.get('user') %}
    <p><a href="{{ url_for('logout') }}">Cerrar sesión</a></p>
  {% else %}
    <p><a href="{{ url_for('login') }}">Iniciar sesión</a></p>
  {% endif %}

  {% with m=get_flashed_messages(with_categories=true) %}
    {% if m %}<div class="alert alert-{{ m[0][0] }}">{{ m[0][1] }}</div>{% endif %}
  {% endwith %}

  <form method="POST" enctype="multipart/form-data" class="border p-4 rounded">
    <!-- Tipo de planilla -->
    <div class="mb-3">
      <label class="form-label">Tipo de planilla</label>
      <select id="mapping_choice" name="mapping_choice"
              class="form-select" onchange="toggleXml()">
        {% for m in mappings %}
        <option value="{{ m|lower }}" {% if m|lower == selected %}selected{% endif %}>{{ m }}</option>
        {% endfor %}
      </select>
    </div>

    <!-- Excel -->
    <div class="mb-3">
      <label class="form-label">Archivo Excel (.xlsx)</label>
      <input class="form-control" type="file" name="file_xlsx" accept=".xlsx" multiple required>
    </div>

    <!-- XML (solo Simbiu) -->
    <div class="mb-3" id="xml_div" style="display:none">
      <label class="form-label">Archivo XML</label>
      <input class="form-control" type="file" name="file_xml" accept=".xml">
    </div>

    <button class="btn btn-primary">Continuar</button>
  </form>
</div>

<script>
function toggleXml(){
  const sel = document.getElementById('mapping_choice');
  document.getElementById('xml_div').style.display =
       sel.value.toLowerCase() === 'simbiu' ? 'block' : 'none';
}
toggleXml();
</script>
"""

TPL_MAPPINGS = """
<!doctype html>
<title>Administrar tipos</title>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css">

<div class="container py-4">
  <a href="{{ url_for('home') }}" class="btn btn-link mb-3">← Menú</a>
  <h2 class="mb-4">Tipos de planilla</h2>
  {% if session.get('user') %}
    <p><a href="{{ url_for('logout') }}">Cerrar sesión</a></p>
  {% else %}
    <p><a href="{{ url_for('login') }}">Iniciar sesión</a></p>
  {% endif %}

  {% with m=get_flashed_messages(with_categories=true) %}
    {% if m %}<div class="alert alert-{{ m[0][0] }}">{{ m[0][1] }}</div>{% endif %}
  {% endwith %}

  <ul class="list-group mb-4">
  {% for m in mappings %}
    <li class="list-group-item">{{ m }}</li>
  {% endfor %}
  </ul>

  <form method="POST" class="mb-4">
    <h5>Crear nuevo tipo</h5>
    <div class="input-group mb-3">
      <input class="form-control" name="new_name" placeholder="Nombre" required>
      <button class="btn btn-success" name="action" value="create">Crear</button>
    </div>

    <h5>Clonar tipo</h5>
    <div class="input-group mb-3">
      <select class="form-select" name="clone_src">
        {% for m in mappings %}<option value="{{ m }}">{{ m }}</option>{% endfor %}
      </select>
      <input class="form-control" name="clone_dest" placeholder="Nuevo nombre" required>
      <button class="btn btn-primary" name="action" value="clone">Clonar</button>
    </div>

    <h5>Eliminar tipo</h5>
    <div class="input-group">
      <select class="form-select" name="delete_name">
        {% for m in mappings if m not in ['lp','simbiu'] %}<option value="{{ m }}">{{ m }}</option>{% endfor %}
      </select>
      <button class="btn btn-danger" name="action" value="delete">Eliminar</button>
    </div>
  </form>
</div>
"""

TPL_MAPPING = """
<!doctype html>
<title>Unificar Excel · {{ mapping_name }}</title>
<link rel="stylesheet"
 href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css">

<div class="container py-4">
  <a href="{{ url_for('home') }}" class="btn btn-link mb-3">← Menú</a>
  <h2 class="mb-4">Configuración: {{ mapping_name }}</h2>
  {% if session.get('user') %}
    <p><a href="{{ url_for('logout') }}">Cerrar sesión</a></p>
  {% else %}
    <p><a href="{{ url_for('login') }}">Iniciar sesión</a></p>
  {% endif %}

  {% with m=get_flashed_messages(with_categories=true) %}
    {% if m %}<div class="alert alert-{{ m[0][0] }}">{{ m[0][1] }}</div>{% endif %}
  {% endwith %}

  <form method="POST">
    {% for sheet, cols in cols_per_sheet.items() %}
      <h4 class="mt-4">{{ sheet }}</h4>
      <table class="table table-sm align-middle">
        <thead><tr><th>Columna original</th><th>Nombre final</th></tr></thead>
        <tbody>
        {% for col in cols %}
          {% set canon = canonicalize(col) %}
          {% set pre = existing_mapping.get(col, existing_mapping.get(canon, canon)) %}
          <tr>
            <td>{{ col }}</td>
            <td><input class="form-control form-control-sm" name="{{ col }}" value="{{ pre }}"></td>
          </tr>
        {% endfor %}
        </tbody>
      </table>
    {% endfor %}
    <div class="d-flex gap-3">
      <button class="btn btn-primary" name="action" value="save">Guardar configuración</button>
      <button class="btn btn-success" name="action" value="unify">Unificar y descargar</button>
      <a href="{{ url_for('home') }}" class="btn btn-secondary">Cancelar</a>
    </div>
  </form>
</div>
"""

TPL_HISTORIAL = """
<!doctype html>
<title>Historial de planillas</title>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css">

<div class="container py-4">
  <a href="{{ url_for('home') }}" class="btn btn-link mb-3">← Menú</a>
  <h2 class="mb-4">Historial de planillas</h2>
  <table class="table table-bordered">
    <thead><tr><th>Fecha</th><th>Tipo</th><th>Archivo</th></tr></thead>
    <tbody>
    {% for r in records %}
      <tr>
        <td>{{ r.fecha }}</td>
        <td>{{ r.mapping }}</td>
        <td><a href="{{ url_for('download_processed', filename=r.archivo) }}">Descargar</a></td>
      </tr>
    {% endfor %}
    </tbody>
  </table>
</div>
"""

# ─────────────────────────── Lanzador ────────────────────────────────
if __name__ == "__main__":
    app.run(debug=True, port=5000)
